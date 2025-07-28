from flask import Flask, render_template, request, redirect, url_for
import pandas as pd
from pathlib import Path
import shutil
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        panel_file = request.files['panel_file']
        office_file = request.files['office_file']
        intervalo_data = request.form['date_range']

        os.makedirs('uploads', exist_ok=True)

        if panel_file and office_file:
            panel_path = Path('uploads') / panel_file.filename
            office_path = Path('uploads') / office_file.filename

            panel_file.save(panel_path)
            office_file.save(office_path)

            # Passe os caminhos reais para a função
            process_files(str(panel_path), str(office_path), intervalo_data)
            return redirect(url_for('index'))

    return render_template('index.html')

def process_files(panel_file, office_file, intervalo_data):
    df = pd.read_excel(panel_file, sheet_name="Sheet1", header=1, engine="openpyxl")
    df.columns = df.columns.str.strip()
    
    office = pd.read_excel(office_file, engine="openpyxl")
    office.columns = office.columns.str.strip()
    mat2nome = dict(zip(office["Matrícula"], office["Nome"]))
    
    pos_aq = df.columns.get_loc("Usuário do Encerramento")
    df.insert(pos_aq + 1, "Nome do Técnico", "")
    df.insert(pos_aq + 2, "Equipe", "")
    
    df["Nome do Técnico"] = df["Usuário do Encerramento"].map(lambda mat: mat2nome.get(mat, "Técnico Operações"))
    df["Equipe"] = df["Usuário do Encerramento"].map(lambda mat: "Recolhimento" if mat in mat2nome else "Técnica")
    
    # Garante que a pasta existe antes de salvar/copiar
    relatorio_dir = Path('Relatório Controladoria')
    relatorio_dir.mkdir(exist_ok=True)
    
    # Crie e salve o arquivo de saída principal
    output_file = relatorio_dir / "Painel_de_Servicos_modificado.xlsx"
    df.to_excel(output_file, index=False)

    # Crie o arquivo final para download
    summary_file = relatorio_dir / "RECOLHIMENTO CONTROLADORIA.xlsx"
    shutil.copy(output_file, summary_file)
    
    criterios = [
        ("Recolhimento", "Fechada Produtiva", "1 Tentativa", ["ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO", "ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO AGENDADO"]),
        ("Recolhimento", "Fechada Improdutiva", "1 Tentativa", ["ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO", "ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO AGENDADO"]),
        ("Técnica", "Fechada Produtiva", "1 Tentativa", ["ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO", "ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO AGENDADO"]),
        ("Técnica", "Fechada Improdutiva", "1 Tentativa", ["ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO", "ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO AGENDADO"]),
        ("Recolhimento", "Fechada Produtiva", "Revisita", ["ESTOQUE - REVISITA DE RECOLHIMENTO EM COMODATO"]),
        ("Recolhimento", "Fechada Improdutiva", "Revisita", ["ESTOQUE - REVISITA DE RECOLHIMENTO EM COMODATO"]),
        ("Técnica", "Fechada Produtiva", "Revisita", ["ESTOQUE - REVISITA DE RECOLHIMENTO EM COMODATO"]),
        ("Técnica", "Fechada Improdutiva", "Revisita", ["ESTOQUE - REVISITA DE RECOLHIMENTO EM COMODATO"])
    ]
    
    tabela = []
    
    for equipe, status, solicitacao, tipos_servico in criterios:
        count = df[(df["Equipe"] == equipe) & (df["Status"] == status) & (df["Tipo de Serviço"].isin(tipos_servico))].shape[0]
        
        tabela.append({
            "Data": intervalo_data,
            "Equipe": equipe,
            "Tipo": "PRODUTIVA" if "Produtiva" in status else "IMPRODUTIVA",
            "Solicitação": solicitacao,
            "Quantidade": count
        })
    
    resultado_df = pd.DataFrame(tabela)
    total = resultado_df["Quantidade"].sum()
    resultado_df.loc[len(resultado_df)] = {
        "Data": "",
        "Equipe": "",
        "Tipo": "",
        "Solicitação": "Total",
        "Quantidade": total
    }
    
    wb = load_workbook(summary_file)
    ws = wb.create_sheet("PRODUTIVIDADE")
    
    for row in dataframe_to_rows(resultado_df, index=False, header=True):
        ws.append(row)
    
    wb.save(summary_file)
    print("Arquivo gerado:", output_file.exists(), output_file)
    return str(summary_file.resolve())  # Retorne o caminho absoluto do arquivo correto

if __name__ == '__main__':
    app.run(debug=True)