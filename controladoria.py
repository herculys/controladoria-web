import pandas as pd
from pathlib import Path
import shutil
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.fill import ColorChoice, PatternFillProperties
from openpyxl.drawing.colors import RGBPercent
from io import BytesIO

def process_files(panel_file, office_file):
    # Leia diretamente dos objetos FileStorage
    df = pd.read_excel(panel_file, sheet_name="Sheet1", header=1, engine="openpyxl")
    df.columns = df.columns.str.strip()
    office = pd.read_excel(office_file, engine="openpyxl")
    office.columns = office.columns.str.strip()
    mat2nome = dict(zip(office["Matrícula"], office["Nome"]))

    # Calcular o intervalo de datas automaticamente
    df["Data/Hora Encerramento"] = pd.to_datetime(df["Data/Hora Encerramento"], errors='coerce')
    data_min = df["Data/Hora Encerramento"].min()
    data_max = df["Data/Hora Encerramento"].max()
    
    # Formatar o intervalo de datas
    if pd.notna(data_min) and pd.notna(data_max):
        if data_min.date() == data_max.date():
            intervalo_data = data_min.strftime("%d/%m/%Y")
        else:
            intervalo_data = f"{data_min.strftime('%d/%m')} à {data_max.strftime('%d/%m/%Y')}"
    else:
        intervalo_data = "Data não disponível"

    pos_aq = df.columns.get_loc("Usuário do Encerramento")
    df.insert(pos_aq + 1, "Nome do Técnico", "")
    df.insert(pos_aq + 2, "Equipe", "")
    df["Nome do Técnico"] = df["Usuário do Encerramento"].map(lambda mat: mat2nome.get(mat, "Técnico Operações"))
    df["Equipe"] = df["Usuário do Encerramento"].map(lambda mat: "Recolhimento" if mat in mat2nome else "Técnica")

    # Gere o arquivo Excel na memória
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name="BASE")
    output.seek(0)

    # Adicione a aba PRODUTIVIDADE
    wb = load_workbook(output)
    ws = wb.create_sheet("PRODUTIVIDADE", 0)  # Criar como primeira planilha (índice 0)
    
    criterios = [
        ("Recolhimento", "Fechada Produtiva", "1 Tentativa", ["ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO", "ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO AGENDADO"]),
        ("Recolhimento", "Fechada Improdutiva", "1 Tentativa", ["ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO", "ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO AGENDADO"]),
        ("Técnica", "Fechada Produtiva", "1 Tentativa", ["ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO", "ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO AGENDADO"]),
        ("Técnica", "Fechada Improdutiva", "1 Tentativa", ["ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO", "ESTOQUE - RECOLHIMENTO DE EQUIPAMENTO COMODATO AGENDADO"]),
        ("Recolhimento", "Fechada Produtiva", "Revisita", ["ESTOQUE - REVISITA DE RECOLHIMENTO EM COMODATO"]),
        ("Recolhimento", "Fechada Improdutiva", "Revisita", ["ESTOQUE - REVISITA DE RECOLHIMENTO EM COMODATO"]),
        ("Técnica", "Fechada Produtiva", "Revisita", ["ESTOQUE - REVISITA DE RECOLHIMENTO EM COMODATO"]),
        ("Técnica", "Fechada Improdutiva", "Revisita", ["ESTOQUE - REVISITA DE RECOLHIMENTO EM COMODATO"])
    ]
    
    tabela = []
    
    for equipe, status, solicitacao, tipos_servico in criterios:
        count = df[(df["Equipe"] == equipe) & (df["Status"] == status) & (df["Tipo de Serviço"].isin(tipos_servico))].shape[0]
        
        tabela.append({
            "Data": intervalo_data,
            "Equipe": equipe,
            "Tipo": "PRODUTIVA" if "Produtiva" in status else "IMPRODUTIVA",
            "Solicitação": solicitacao,
            "Quantidade": count
        })
    
    resultado_df = pd.DataFrame(tabela)
    total = resultado_df["Quantidade"].sum()
    resultado_df.loc[len(resultado_df)] = {
        "Data": "",
        "Equipe": "",
        "Tipo": "",
        "Solicitação": "Total",
        "Quantidade": total
    }
    
    for row in dataframe_to_rows(resultado_df, index=False, header=True):
        ws.append(row)
    
    # Criar os 4 gráficos de pizza usando dados diretamente da tabela
    chart_positions = [
        ("Recolhimento", "1 Tentativa", "H2", "Recolhimento - 1ª Tentativa"),
        ("Técnica", "1 Tentativa", "P2", "Técnica - 1ª Tentativa"),
        ("Recolhimento", "Revisita", "H18", "Recolhimento - Revisita"),
        ("Técnica", "Revisita", "P18", "Técnica - Revisita")
    ]
    
    # Encontrar as linhas correspondentes na tabela para cada gráfico
    for i, (equipe, solicitacao, chart_pos, title) in enumerate(chart_positions):
        # Encontrar as linhas da tabela que correspondem a este gráfico
        produtiva_row = None
        improdutiva_row = None
        
        # Procurar nas linhas da tabela (começando da linha 2, já que linha 1 é cabeçalho)
        for row_num in range(2, ws.max_row + 1):
            equipe_cell = ws.cell(row=row_num, column=2).value  # Coluna B - Equipe
            tipo_cell = ws.cell(row=row_num, column=3).value    # Coluna C - Tipo
            solicitacao_cell = ws.cell(row=row_num, column=4).value  # Coluna D - Solicitação
            
            if equipe_cell == equipe and solicitacao_cell == solicitacao:
                if tipo_cell == "PRODUTIVA":
                    produtiva_row = row_num
                elif tipo_cell == "IMPRODUTIVA":
                    improdutiva_row = row_num
        
        # Criar gráfico apenas se encontramos ambas as linhas
        if produtiva_row and improdutiva_row:
            # Criar gráfico de pizza
            chart = PieChart()
            chart.title = title
            
            # Usar dados diretamente da tabela
            # Dados: coluna E (Quantidade) das linhas encontradas
            data = Reference(ws, min_col=5, min_row=produtiva_row, max_row=improdutiva_row)
            # Categorias: coluna C (Tipo) das linhas encontradas  
            categories = Reference(ws, min_col=3, min_row=produtiva_row, max_row=improdutiva_row)
            
            chart.add_data(data)
            chart.set_categories(categories)
            
            # Configurar tamanho do gráfico
            chart.width = 10     # Largura 10cm
            chart.height = 6     # Altura 6cm
            
            # Configurar rótulos com quantidade e porcentagem (sem nome da série)
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showCatName = False  # Remove o nome da categoria
            chart.dataLabels.showSerName = False  # Remove o nome da série
            chart.dataLabels.showVal = True
            chart.dataLabels.showPercent = True
            
            # Configurar cores das fatias
            # PRODUTIVA = Rosa forte, IMPRODUTIVA = Azul escuro
            if chart.series:
                serie = chart.series[0]
                if len(serie.dPt) == 0:
                    # Adicionar pontos de dados se não existirem
                    from openpyxl.chart.data_source import NumDataSource
                    from openpyxl.chart.series import DataPoint
                    
                    # Ponto 0: PRODUTIVA (Rosa forte)
                    pt0 = DataPoint(idx=0)
                    pt0.graphicalProperties.solidFill = "FF3399"  # Rosa forte
                    serie.dPt.append(pt0)
                    
                    # Ponto 1: IMPRODUTIVA (Azul escuro)
                    pt1 = DataPoint(idx=1)
                    pt1.graphicalProperties.solidFill = "0000C6"  # Azul escuro
                    serie.dPt.append(pt1)
            
            # Posicionar o gráfico
            ws.add_chart(chart, chart_pos)
    
    # Salve novamente na memória
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output